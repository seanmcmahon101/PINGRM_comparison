import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

def select_file(file_description="Excel files", extensions="*.xls *.xlsx"):
    """Opens a file dialog to select a file and returns the file path."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(title=f"Select file for {file_description}", filetypes=[(file_description, extensions)])
    return file_path

def process_trim_and_filter():
    pingrm_file = select_file("PINGRM Schedule files", "*.csv *.xlsx")
    if not pingrm_file:
        print("File selection cancelled.")
        return None

    file_extension = os.path.splitext(pingrm_file)[1].lower()  # Normalize file extension
    print(f"File selected: {pingrm_file} with extension: {file_extension}")

    if file_extension == '.csv':
        df = pd.read_csv(pingrm_file, sep=';', header=None, dtype=str, skiprows=1)
    elif file_extension == '.xlsx':
        df = pd.read_excel(pingrm_file)
    else:
        print(f"Unsupported file format: {file_extension}. Please select a CSV or Excel file.")
        return None

    if file_extension == '.csv':
        df.replace(';;', ';', regex=True, inplace=True)

    processed_data = []
    for index, row in df.iterrows():
        item_number, term_date, order_quantity, reference, year_week = "", "", "", "", ""
        for value in row:
            if pd.notna(value):
                if len(str(value)) == 10 and str(value).startswith('4'):
                    item_number = str(value)
                elif pd.to_datetime(value, errors='coerce') is not pd.NaT and term_date == "":
                    term_date = pd.to_datetime(value).strftime('%d-%m-%Y')
                    year_week = f"{pd.to_datetime(value).year}-{pd.to_datetime(value).isocalendar().week:02d}"
                elif str(value).isnumeric() and order_quantity == "" and term_date != "":
                    order_quantity = str(value)
                elif (len(str(value)) == 8 and str(value).startswith('3')) or str(value) == "forecast":
                    reference = str(value)
        processed_data.append({
            'Item Number': item_number, 'Term Date': term_date, 'Order Quantity': order_quantity,
            'Reference': reference, 'Year Week': year_week
        })
    return pd.DataFrame(processed_data)

def process_import_and_organize():
    codate_file = select_file("CODATE files", "*.xls *.xlsx")
    if not codate_file:
        print("No file selected.")
        return None

    if codate_file.endswith('.xlsx'):
        engine = 'openpyxl'
    elif codate_file.endswith('.xls'):
        engine = 'xlrd'
    else:
        raise ValueError("Unsupported file format")

    df = pd.read_excel(codate_file, engine=engine, usecols=[
        'CustID', 'Customer PONumber', 'CONumber', 'Ln', 
        'Item Number', 'Item Description', 'Cust Item Number', 
        'OrderQty', 'Open Qty', 'PromDlvry'
    ])
    df['PromDlvry'] = pd.to_datetime(df['PromDlvry'], errors='coerce')
    df["Year-Week"] = df['PromDlvry'].dt.strftime('%Y-%U')
    df.sort_values('Year-Week', inplace=True)
    
    df = df[df['CustID'] == 'PINGRM']  # Filter 'CustID' for 'PINGRM'
    
    return df

def export_data():
    df1 = process_trim_and_filter()
    df2 = process_import_and_organize()

    if df1 is None or df2 is None:
        print("Data processing was cancelled for one or both datasets.")
        return

    workbook = Workbook()
    ws1 = workbook.active
    ws1.title = "PINGRM"

    for r_idx, row in enumerate(df1.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=value)
    for c_idx, col in enumerate(df1.columns, start=1):
        ws1.cell(row=1, column=c_idx, value=col)

    ws2 = workbook.create_sheet("HF")
    for r_idx, row in enumerate(df2.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    for c_idx, col in enumerate(df2.columns, start=1):
        ws2.cell(row=1, column=c_idx, value=col)

    add_summary_sheet(workbook)

    output_filename = 'Processed_Data.xlsx'
    counter = 1
    while os.path.exists(output_filename):
        output_filename = f'Processed_Data_{counter}.xlsx'
        counter += 1

    workbook.save(output_filename)
    print(f"Data exported to {output_filename}")
    return workbook

def add_summary_sheet(workbook):
    ws_pingrm = workbook["PINGRM"]
    ws_hf = workbook["HF"]
    ws_summary = workbook.create_sheet("Summary")
    ws_summary.append(["Item Number", "Reference", "Year Week (PINGRM)", "Year-Week (HF)", "Date Change Detected"])

    # Convert sheets to DataFrames
    df_pingrm = pd.DataFrame([cell.value for cell in row] for row in ws_pingrm.iter_rows(min_row=2))
    df_hf = pd.DataFrame([cell.value for cell in row] for row in ws_hf.iter_rows(min_row=2))
    df_pingrm.columns = ["Item Number", "Term Date", "Order Quantity", "Reference", "Year Week"]
    df_hf.columns = ["CustID", "Customer PONumber", "CONumber", "Ln", "Item Number", "Item Description", "Cust Item Number", "OrderQty", "Open Qty", "PromDlvry", "Year-Week"]

    # Sort 'HF' DataFrame by 'CONumber' from newest to oldest if it is a date or numeric
    df_hf['CONumber'] = pd.to_numeric(df_hf['CONumber'], errors='coerce')  # Ensure 'CONumber' is treated as numeric for sorting
    df_hf = df_hf.sort_values(by='CONumber', ascending=False)

    # Append data to summary sheet and check for changes
    for index, row in df_pingrm.iterrows():
        item_number = row['Item Number']
        reference = row['Reference']
        year_week_pingrm = row['Year Week']
        matching_rows = df_hf[df_hf['Customer PONumber'] == reference]
        year_week_hf = matching_rows['Year-Week'].iloc[0] if not matching_rows.empty else "No match found"
        date_change = "No" if year_week_pingrm == year_week_hf else "Yes" if not matching_rows.empty else "Reference not found"
        ws_summary.append([item_number, reference, year_week_pingrm, year_week_hf, date_change])

    # Convert the Summary data to a table for better formatting and usability
    tab = Table(displayName="SummaryTable", ref=ws_summary.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws_summary.add_table(tab)

# To run the entire process, simply call export_data()
export_data()

